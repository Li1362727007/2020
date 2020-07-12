import openpyxl
import  requests
def read_data(filename,sheetname):
    wb=openpyxl.load_workbook(filename)
    sheet=wb[sheetname]
    max_row=sheet.max_row
    max_column=sheet.max_column
    all_cases=[]
    for i in range(2,max_row+1):
        cases=[]
        for j in range(1,max_column+1):
            case=sheet.cell(row=i,column=j).value
            cases.append(case)
        all_cases.append(cases)
    #print(all_cases)
    return all_cases
def requests_login(url,data,token=None):
    header = {'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json', 'Authorization': token}
    response=requests.post(url,data,headers=header)
    result=response.json()
    return result


def write_data(filename,sheetname,row,column,final_result):
    wb=openpyxl.load_workbook(filename)
    sheet=wb[sheetname]
    sheet.cell(row,column).value=filename
    wb.save(filename)

url = "http://120.78.128.25:8766/futureloan/member/login"
payload = '{"mobile_phone": "18855123497","pwd": "lemon123456"}'
login_result=requests_login(url,payload)
print(login_result)
token='Bearer '+login_result['data']['token_info']['token']
all_cases=read_data('test_jiekou.xlsx','recharge')
for case in all_cases:
    case_url=case[5]
    case_data=case[6]
    #case_data=eval(case_data)
    expect_result=case[7]
    real_result=requests_login(case_url,case_data,token=token)
    print(real_result)

