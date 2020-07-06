import requests
import openpyxl
url = "http://120.78.128.25:8766/futureloan/member/login"
payload = '{"mobile_phone": "18855123497","pwd": "lemon123456"}'
headers = {'Content-Type': 'application/json','X-Lemonban-Media-Type': 'lemonban.v2'}
header={'Content-Type': 'application/json','X-Lemonban-Media-Type':'lemonban.v2','Authorization':'Bearer eyJhbGciOiJIUzUxMiJ9.eyJtZW1iZXJfaWQiOjk0LCJleHAiOjE1OTE1NzM3OTJ9.pZYq0wInSMctOLPqFe_T7EEBnteUzBfSlDjkoBghIgdUqpNYCaqxBUoWVOfKhSYbFgYlDKl7L9fJgszDTdtnEg'}
response=requests.post(url,payload,headers=headers)
print(response.json())
def read_data(filename,sheetname):
    wb=openpyxl.load_workbook(filename)
    sheet=wb[sheetname]
    max_row=sheet.max_row
    list_dict=[]
    for i in range(2,max_row+1):
        case=dict(
            case_id=sheet.cell(row=i,column=1).value,
            url=sheet.cell(row=i,column=6).value,
            data=sheet.cell(row=i,column=7).value,
            expect_result=sheet.cell(row=i,column=8).value
        )
        list_dict.append(case)
    return list_dict

def post_func(post_url,post_data,headers=header):
    response=requests.post(url=post_url,data=post_data,headers=header)
    result=response.json()
    return result
def write_result(filename,sheetname,row,column,final_result):
    wb=openpyxl.load_workbook(filename)
    sheet=wb[sheetname]
    sheet.cell(row,column).value=final_result
    wb.save(filename)
cases=read_data('test_case.xlsx','recharge')
for case in cases:
    case_id=case['case_id']
    url=case.get('url')
    data=case.get('data')
    #data=data.replace("101","94")
   # data=eval(data)
    expect_result=case.get('expect_result')
    expect_result=expect_result.replace('null','None')
    expect_result=eval(expect_result)
    real_result=post_func(post_url=url,post_data=data,headers=header)
    real_code=real_result.get('code')
    real_msg=real_result.get('msg')
    expect_code=expect_result.get('code')
    print('真实的执行结果是：{}'.format(real_msg))
    if real_code==expect_code:
        print('第{}输入结果正确'.format(case_id))
        result="Passed"
    else:
        print('第{}输入的结果错误'.format(case_id))
        result='Failed'
    print('*'*20)
    write_result('test_case.xlsx','recharge',case_id+1,10,result)